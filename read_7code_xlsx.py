# 7CODE 다이어그램 코드별 정리 엑셀 파일을 읽어서 데이터를 추출하고 확인하기 위한 임시 스크립트입니다.
import openpyxl
import os

xlsx_name = '7CODE 다이어그램 코드별 정리.xlsx'
if not os.path.exists(xlsx_name):
    print(f"파일이 존재하지 않습니다: {xlsx_name}")
else:
    wb = openpyxl.load_workbook(xlsx_name, data_only=True)
    with open('7code_extracted.txt', 'w', encoding='utf-8') as f_out:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            f_out.write(f"=== 시트: {sheet_name} ===\n")
            f_out.write(f"행 수: {ws.max_row}, 열 수: {ws.max_column}\n")
            for r_idx in range(1, ws.max_row + 1):
                row_vals = []
                for c_idx in range(1, ws.max_column + 1):
                    val = ws.cell(row=r_idx, column=c_idx).value
                    row_vals.append(val)
                # 행이 전부 None이 아니면 출력
                if any(v is not None for v in row_vals):
                    safe_row_vals = [str(v) if v is not None else 'None' for v in row_vals]
                    f_out.write(f"Row {r_idx}: {', '.join(safe_row_vals)}\n")

