# 엑셀 34명을 JSON에서 branchId + 나이 + 성별로 매칭하여 전체 데이터 추출
import json
import openpyxl

EXCEL = "춘천교육청회원데이터_분석완료_2026-06-24.xlsx"
EVT = "EVT_춘천_260605_7UNA"

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb[wb.sheetnames[0]]
headers = {}
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=c).value
    if val:
        headers[str(val).strip()] = c

print("엑셀 헤더:", headers)

# 엑셀에서 매칭 키 생성 (지점ID + 나이)
excel_members = []
for r in range(2, ws.max_row + 1):
    name = str(ws.cell(row=r, column=headers["이름"]).value or "").strip()
    age = ws.cell(row=r, column=headers["나이"]).value
    gender = str(ws.cell(row=r, column=headers["성별"]).value or "").strip()
    branch = str(ws.cell(row=r, column=headers["지점"]).value or "").strip()
    if name:
        excel_members.append({"name": name, "age": int(age) if age else None, "gender": gender, "branch": branch})
        
print(f"엑셀 회원: {len(excel_members)}명")
for m in excel_members[:5]:
    print(f"  {m['name']} / {m['age']}세 / {m['gender']} / branch={m['branch'][:10]}...")

# JSON 로드
with open("cleaned_members_v2.json", "r", encoding="utf-8") as f:
    all_members = json.load(f)

# 합동코드 필터
evt_members = [m for m in all_members if m.get("eventCode") == EVT and m.get("report")]
print(f"\nJSON 합동코드 매칭: {len(evt_members)}건")

# JSON에서도 매칭 키 생성 (branchId + 나이)
for jm in evt_members[:3]:
    report = jm.get("report", {})
    ui = report.get("userInfo", {})
    jname = (ui.get("name") or jm.get("name") or "").strip()
    jage = ui.get("age")
    jgender = (ui.get("gender") or "").strip()
    jbranch = jm.get("branchId", "")
    print(f"  JSON: {jname} / {jage}세 / {jgender} / branch={jbranch[:10]}...")

# 매칭 전략 1: branchId + 나이 조합
print("\n=== 매칭 시도: branchId + 나이 ===")
matched = {}
for em in excel_members:
    key = f"{em['branch']}_{em['age']}"
    for jm in evt_members:
        report = jm.get("report", {})
        ui = report.get("userInfo", {})
        jage = ui.get("age")
        jbranch = jm.get("branchId", "")
        jkey = f"{jbranch}_{jage}"
        if key == jkey and em["name"] not in matched:
            matched[em["name"]] = jm
            break

print(f"branchId+나이 매칭: {len(matched)}명 / {len(excel_members)}명")

# 매칭 전략 2: 나이 + 성별
if len(matched) < len(excel_members):
    print("\n=== 매칭 시도: 나이 + 성별 ===")
    gender_map = {"남": "male", "여": "female"}
    for em in excel_members:
        if em["name"] in matched:
            continue
        em_gender = gender_map.get(em["gender"], "")
        for jm in evt_members:
            if id(jm) in [id(v) for v in matched.values()]:
                continue
            report = jm.get("report", {})
            ui = report.get("userInfo", {})
            jage = ui.get("age")
            jgender = (ui.get("gender") or "").strip().lower()
            if em["age"] == jage and em_gender == jgender and em["name"] not in matched:
                matched[em["name"]] = jm
                break
    print(f"누적 매칭: {len(matched)}명 / {len(excel_members)}명")

# 매칭된 데이터에서 마음나이, 3Body, 7Code 확인
print("\n=== 매칭 데이터 확인 ===")
found_mind = 0
found_3body = 0
found_7code = 0
for name, jm in list(matched.items())[:5]:
    report = jm.get("report", {})
    mind_age = report.get("mindAge")
    tba = report.get("threeBodyAnalysis", {})
    sca = report.get("sevenCodeAnalysis", {})
    
    body_data = tba.get("body", {})
    body_score = body_data.get("score") if isinstance(body_data, dict) else body_data
    brain_data = tba.get("brain", {})
    brain_score = brain_data.get("score") if isinstance(brain_data, dict) else brain_data
    mind_data = tba.get("mind", {})
    mind_score = mind_data.get("score") if isinstance(mind_data, dict) else mind_data
    
    print(f"  {name}: mindAge={mind_age}, 3Body=({body_score}/{brain_score}/{mind_score}), 7Code keys={list(sca.keys())[:3]}")

for name, jm in matched.items():
    report = jm.get("report", {})
    if report.get("mindAge") is not None:
        found_mind += 1
    if report.get("threeBodyAnalysis"):
        found_3body += 1
    if report.get("sevenCodeAnalysis"):
        found_7code += 1

print(f"\n마음나이 있는 회원: {found_mind}/{len(matched)}")
print(f"3Body 있는 회원:    {found_3body}/{len(matched)}")
print(f"7Code 있는 회원:    {found_7code}/{len(matched)}")

# 미매칭 목록
unmatched = [em["name"] for em in excel_members if em["name"] not in matched]
print(f"\n미매칭: {len(unmatched)}명 - {unmatched}")
