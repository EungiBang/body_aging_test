# 이름 매칭 디버그 스크립트
import json
import openpyxl

# 엑셀에서 이름 추출
wb = openpyxl.load_workbook("춘천교육청회원데이터_분석완료_2026-06-24.xlsx", data_only=True)
ws = wb[wb.sheetnames[0]]
excel_names = set()
for r in range(2, ws.max_row + 1):
    name = ws.cell(row=r, column=1).value
    if name:
        n = str(name).strip()
        excel_names.add(n)
        if r <= 5:
            print(f"엑셀 이름: '{n}' (type={type(name)}, repr={repr(n)})")

# JSON에서 합동코드 멤버 이름 추출
with open("cleaned_members_v2.json", "r", encoding="utf-8") as f:
    all_members = json.load(f)

evt = [m for m in all_members if m.get("eventCode") == "EVT_춘천_260605_7UNA" and m.get("report")]
json_names = set()
for m in evt:
    report = m.get("report", {})
    ui = report.get("userInfo", {})
    name = (ui.get("name") or m.get("name") or "").strip()
    json_names.add(name)

print(f"\n엑셀 이름 샘플: {sorted(list(excel_names))[:5]}")
print(f"JSON 이름 샘플: {sorted(list(json_names))[:5]}")

# 직접 비교
print(f"\n엑셀 이름 수: {len(excel_names)}")
print(f"JSON 이름 수: {len(json_names)}")
print(f"교집합: {len(excel_names & json_names)}")
print(f"엑셀에만: {excel_names - json_names}")
print(f"JSON에만: {json_names - excel_names}")

# 바이트 레벨 비교
excel_first = sorted(list(excel_names))[0]
json_first = sorted(list(json_names))[0]
print(f"\n엑셀 첫 이름 bytes: {excel_first.encode('utf-8')}")
print(f"JSON 첫 이름 bytes: {json_first.encode('utf-8')}")
