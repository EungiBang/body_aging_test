# openpyxl로 직접 읽어서 인코딩 깨짐 없이 확인
import openpyxl

wb = openpyxl.load_workbook('3바디 설치현황0419.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== 시트: {sheet_name} ===")
    print(f"행 수: {ws.max_row}, 열 수: {ws.max_column}")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [(cell.value, cell.row, cell.column) for cell in row if cell.value is not None]
        if vals:
            print([f"({v[1]},{v[2]})={v[0]}" for v in vals])
