# 춘천교육청 엑셀 파일 구조 확인 스크립트
import openpyxl

wb = openpyxl.load_workbook("춘천교육청회원데이터_분석완료_2026-06-24.xlsx", data_only=True)
print(f"시트 목록: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== 시트: {sheet_name} ===")
    print(f"행: {ws.max_row}, 열: {ws.max_column}")
    
    # 헤더 출력
    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        headers.append(str(val) if val else "")
    print(f"헤더: {headers}")
    
    # 처음 5행 샘플
    for r in range(2, min(ws.max_row + 1, 7)):
        row_data = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            row_data.append(str(val)[:30] if val else "")
        print(f"  행{r}: {row_data}")
    
    # 전체 이름 목록
    print(f"\n  전체 이름 목록:")
    names = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if name:
            names.append(str(name).strip())
    print(f"  총 {len(names)}명: {names}")
