# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "지역_지점_등록"

# 스타일 정의
header_font = Font(name='맑은 고딕', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
region_font = Font(name='맑은 고딕', bold=True, size=11, color='1F3864')
region_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
branch_font = Font(name='맑은 고딕', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 헤더
headers = ['지역', '지점', '할당량']
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 10

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# === 전체 지역-지점 데이터 (엑셀에서 추출한 정확한 데이터) ===
data = {
    "본사": ["신입경영홍보실"],
    "서울강북1": ["공덕", "광화문", "연신내", "왕십리", "홍제"],
    "서울강북2": ["구리", "구의", "성북", "수유", "의정부", "창동"],
    "서울강북3": ["공릉", "마들", "상계"],
    "서울강남1": ["고덕", "둔촌", "서초", "잠실", "강남"],
    "서울강남2": ["강서", "신도림", "양천구", "여의도", "이수"],
    "경기북부": ["김포", "마두", "일산", "파주", "화정"],
    "경기남부1": ["고잔", "과천", "산본", "안양1번가", "평촌"],
    "경기남부2": ["미금", "서현", "수지", "야탑", "이천"],
    "경기남부3": ["동탄", "영통", "오산", "천천동", "평택"],
    "인천": ["검단", "논현", "부평", "상동", "송도", "주안"],
    "강원": ["강릉", "동해", "원주", "춘천"],
    "대전": ["둔산", "서대전", "송촌", "월평"],
    "충북": ["가경", "금천", "율량", "제천"],
    "충남": ["서산", "세종아름", "아산", "불당", "충무"],
    "광주전남": ["동광주", "서구", "순천", "여수", "진월", "첨단"],
    "전북": ["군산", "익산", "인후", "정읍", "효자"],
    "대구": ["방촌", "범어", "상인", "성서", "수성", "시지", "월배", "침산"],
    "경북": ["경산", "경주", "구미", "김천", "대이", "두호", "문경", "상주", "안동", "영주"],
    "경남": ["거제", "내외", "마산", "명곡", "진주", "진해", "창원"],
    "울산": ["대송", "북구", "울산", "중구"],
    "부산": ["금정", "기장", "다대", "대연", "동래", "사직", "양산", "양정", "영도", "온천장", "하단"],
    "제주": ["노형", "서귀포", "일도"],
    "단무도": ["단무도대구", "단무도부산", "단무도부천", "단무도분당", "단무도서울", "단무도일산", "단무도창원", "단무도천안", "단무도평촌"],
}

row = 2
for region, branches in data.items():
    for i, branch in enumerate(branches):
        cell_a = ws.cell(row=row, column=1, value=region if i == 0 else "")
        cell_b = ws.cell(row=row, column=2, value=branch)
        cell_c = ws.cell(row=row, column=3, value=2)  # 기본 할당량 2대

        if i == 0:
            cell_a.font = region_font
            cell_a.fill = region_fill
        cell_a.border = thin_border
        cell_a.alignment = Alignment(vertical='center')

        cell_b.font = branch_font
        cell_b.border = thin_border
        cell_b.alignment = Alignment(vertical='center')

        cell_c.font = branch_font
        cell_c.border = thin_border
        cell_c.alignment = Alignment(horizontal='center', vertical='center')

        row += 1

# 지역별 병합 (가독성 향상)
row = 2
for region, branches in data.items():
    count = len(branches)
    if count > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row + count - 1, end_column=1)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    row += count

output_path = 'BTC_지점_등록_데이터_v4.xlsx'
wb.save(output_path)

# 통계 출력
total_branches = sum(len(b) for b in data.values())
print(f"✅ 파일 생성 완료: {output_path}")
print(f"   총 {len(data)}개 지역, {total_branches}개 지점")
for region, branches in data.items():
    print(f"   📍 {region}: {', '.join(branches)}")
