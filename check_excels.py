# 기존 엑셀 파일들의 시트 정보와 행 수를 분석하는 스크립트
import openpyxl
import os

files = [
    'BTC_코드맵_점검_통계보고서_수정.xlsx'
]

for f_name in files:
    if os.path.exists(f_name):
        try:
            wb = openpyxl.load_workbook(f_name, read_only=True)
            print(f"\n===== Excel: {f_name} =====")
            print("Sheets:", wb.sheetnames)
            for s_name in wb.sheetnames:
                ws = wb[s_name]
                print(f"  Sheet '{s_name}': max_row={ws.max_row}, max_column={ws.max_column}")
        except Exception as e:
            print(f"Error reading {f_name}: {e}")
    else:
        print(f"File not found: {f_name}")
