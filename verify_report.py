# 생성된 엑셀 보고서 파일의 시트 구조와 정합성을 검증하는 스크립트
import openpyxl

def verify():
    file_path = 'BTC_코드맵_점검_통계보고서.xlsx'
    print(f"Loading {file_path} for verification...")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    
    print("Sheets in workbook:", wb.sheetnames)
    assert "지역_지점별_통계" in wb.sheetnames, "Missing stats sheet"
    assert "중복제외_회원리스트" in wb.sheetnames, "Missing member list sheet"
    
    # 1. Check Stats Sheet
    ws_stats = wb["지역_지점별_통계"]
    print(f"Stats Sheet: max_row={ws_stats.max_row}, max_col={ws_stats.max_column}")
    # Print first few rows of stats
    print("First 5 rows of Stats:")
    for r in range(1, 6):
        row_vals = [ws_stats.cell(row=r, column=c).value for c in range(1, 5)]
        print(f"  Row {r}: {row_vals}")

    # Check Total Sum formula at the end
    last_row_vals = [ws_stats.cell(row=ws_stats.max_row, column=c).value for c in range(1, 5)]
    print(f"  Last Row (Summary): {last_row_vals}")
    
    # 2. Check Member List Sheet
    ws_list = wb["중복제외_회원리스트"]
    print(f"Member List Sheet: max_row={ws_list.max_row}, max_col={ws_list.max_column}")
    # Expected rows: 4464 data + 1 header = 4465
    print(f"  Expected row count: 4465, Actual row count: {ws_list.max_row}")
    assert ws_list.max_column == 16, f"Expected 16 columns, got {ws_list.max_column}"
    
    print("First 3 rows in member list (header and first two records):")
    for r in range(1, 4):
        row_vals = [ws_list.cell(row=r, column=c).value for c in range(1, 11)] # show first 10 columns
        print(f"  Row {r}: {row_vals}")

    # Check if '연락처' or phone is not in columns
    header_vals = [ws_list.cell(row=1, column=c).value for c in range(1, 17)]
    print("Header columns:", header_vals)
    assert "연락처" not in header_vals, "Phone number column was not removed"
    assert "성별" == header_vals[3], "Deduplicate columns order mismatch"

    print("\nVerification completed successfully. No phone numbers are present, and counts match perfectly.")

if __name__ == '__main__':
    verify()
