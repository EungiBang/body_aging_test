# 분석 완료된 회원 데이터를 기반으로 중복을 제거한 지역/지점별 통계 및 고유 회원 목록 엑셀 보고서를 생성하는 스크립트
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel_report():
    print("==================================================")
    print("  지점별 코드맵 점검 통계 및 회원 리스트 생성")
    print("==================================================")

    # 1. 데이터 로드
    if not os.path.exists('cleaned_members_v2.json') or not os.path.exists('firebase_dump_v2.json'):
        print("[오류] 필요한 데이터 파일(cleaned_members_v2.json, firebase_dump_v2.json)이 존재하지 않습니다.")
        return

    with open('cleaned_members_v2.json', 'r', encoding='utf-8') as f:
        members = json.load(f)
    
    with open('firebase_dump_v2.json', 'r', encoding='utf-8') as f:
        dump = json.load(f)

    regions = dump.get('regions', {})
    branches = dump.get('branches', {})

    print(f"로드된 전체 회원 레코드: {len(members)}건")
    print(f"로드된 지역: {len(regions)}개, 지점: {len(branches)}개")

    # 2. 분석 완료된 회원 필터링 (report가 존재하고 id가 pending-으로 시작하지 않는 레코드)
    analyzed_members = []
    for m in members:
        m_id = m.get('id', '')
        report = m.get('report')
        if report and not m_id.startswith('pending-'):
            analyzed_members.append(m)

    print(f"분석 완료 회원 (사용자 언급 4,793명 조건): {len(analyzed_members)}건")

    # 3. 중복 제거 처리
    # 사용자의 피드백(휴대폰 번호 누락 많음 및 개인정보 보호)을 반영하여
    # 연락처(phone)를 판정 기준에서 제외하고 (이름, 지점 ID, 나이, 성별)을 기준으로 중복을 판정합니다.
    seen_keys = set()
    unique_members = []
    duplicate_count = 0

    for m in analyzed_members:
        name = m.get('name', '').strip()
        r = m.get('report', {})
        ui = r.get('userInfo', {}) if isinstance(r, dict) else {}
        gender = ui.get('gender', '')
        age = str(ui.get('age', ''))
        branch_id = m.get('branchId', '')

        # (이름, 지점 ID, 나이, 성별)을 고유 판정 키로 지정
        key = (name, branch_id, age, gender)

        if key not in seen_keys:
            seen_keys.add(key)
            unique_members.append(m)
        else:
            duplicate_count += 1

    print(f"중복 제거 결과: 고유 회원 {len(unique_members)}명 (중복 {duplicate_count}건 제외됨)")

    # 4. 통계 집계 (지역별, 지점별)
    # raw_stats: { region_name: { branch_name: { 'raw_count': 0, 'unique_count': 0 } } }
    # 매핑 정보 활용
    stats = {}

    for m in analyzed_members:
        r_id = m.get('regionId', '')
        b_id = m.get('branchId', '')

        # 지역명 구하기
        region_name = regions.get(r_id, {}).get('name', '')
        if not region_name:
            region_name = r_id.replace('region_', '') if r_id else '미지정'
        
        # 지점명 구하기
        branch_name = branches.get(b_id, {}).get('name', '')
        if not branch_name:
            branch_name = '미지정 지점'

        if region_name not in stats:
            stats[region_name] = {}
        if branch_name not in stats[region_name]:
            stats[region_name][branch_name] = {'raw_count': 0, 'unique_members': []}

        stats[region_name][branch_name]['raw_count'] += 1

    # 고유 회원 정보를 기반으로 지점별 고유 점검인원수 추가
    for m in unique_members:
        r_id = m.get('regionId', '')
        b_id = m.get('branchId', '')

        region_name = regions.get(r_id, {}).get('name', '')
        if not region_name:
            region_name = r_id.replace('region_', '') if r_id else '미지정'
        
        branch_name = branches.get(b_id, {}).get('name', '')
        if not branch_name:
            branch_name = '미지정 지점'

        if region_name in stats and branch_name in stats[region_name]:
            stats[region_name][branch_name]['unique_members'].append(m)

    # 5. 엑셀 워크북 생성 및 디자인 설정
    wb = openpyxl.Workbook()
    
    # 시트 1: 지역 및 지점별 통계
    ws_stats = wb.active
    ws_stats.title = "지역_지점별_통계"
    ws_stats.views.sheetView[0].showGridLines = True

    # 시트 2: 중복제외 회원 리스트
    ws_list = wb.create_sheet(title="중복제외_회원리스트")
    ws_list.views.sheetView[0].showGridLines = True

    # 스타일 요소 정의
    font_family = "맑은 고딕"
    
    # 헤더 스타일
    header_font = Font(name=font_family, bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # 신뢰감을 주는 딥 네이비
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 지역 행(그룹 첫 행) 스타일
    region_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    region_font = Font(name=font_family, bold=True, size=10, color="1F497D")

    # 일반 데이터 스타일
    data_font = Font(name=font_family, size=10)
    data_align_left = Alignment(horizontal="left", vertical="center")
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_right = Alignment(horizontal="right", vertical="center")

    # 요약/합계 스타일
    summary_font = Font(name=font_family, bold=True, size=10, color="000000")
    summary_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # 테두리 스타일
    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_bottom_side = Side(style="medium", color="1F497D")
    double_bottom_side = Side(style="double", color="000000")
    
    summary_border = Border(
        left=thin_side, right=thin_side,
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000")
    )

    # --- 시트 1: 통계 생성 ---
    headers_stats = ["지역", "지점", "총 점검 건수 (중복 포함)", "실제 점검 인원 (중복 제외)"]
    ws_stats.row_dimensions[1].height = 28
    
    for col_idx, h in enumerate(headers_stats, 1):
        cell = ws_stats.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 정렬하여 데이터 쓰기
    sorted_regions = sorted(stats.keys())
    row_idx = 2
    
    region_merge_info = [] # (start_row, end_row, region_name)

    for r_name in sorted_regions:
        start_row = row_idx
        sorted_branches = sorted(stats[r_name].keys())
        
        for b_name in sorted_branches:
            ws_stats.row_dimensions[row_idx].height = 20
            
            b_info = stats[r_name][b_name]
            raw_cnt = b_info['raw_count']
            uniq_cnt = len(b_info['unique_members'])

            # 셀에 값 작성
            cell_region = ws_stats.cell(row=row_idx, column=1, value=r_name)
            cell_branch = ws_stats.cell(row=row_idx, column=2, value=b_name)
            cell_raw = ws_stats.cell(row=row_idx, column=3, value=raw_cnt)
            cell_uniq = ws_stats.cell(row=row_idx, column=4, value=uniq_cnt)

            # 포맷 및 스타일 지정
            cell_region.font = data_font
            cell_region.alignment = data_align_center
            cell_region.border = thin_border

            cell_branch.font = data_font
            cell_branch.alignment = data_align_left
            cell_branch.border = thin_border

            cell_raw.font = data_font
            cell_raw.alignment = data_align_right
            cell_raw.number_format = '#,##0'
            cell_raw.border = thin_border

            cell_uniq.font = data_font
            cell_uniq.alignment = data_align_right
            cell_uniq.number_format = '#,##0'
            cell_uniq.border = thin_border

            row_idx += 1
            
        end_row = row_idx - 1
        region_merge_info.append((start_row, end_row, r_name))

    # 지역 병합 및 스타일링
    for start_r, end_r, r_name in region_merge_info:
        if start_r < end_r:
            ws_stats.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
            # 첫 번째 셀만 정렬 설정
            merged_cell = ws_stats.cell(row=start_r, column=1)
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 지역 블록 전체에 통일감 있는 은은한 채우기 효과 적용 (가독성 향상)
        for r in range(start_r, end_r + 1):
            cell = ws_stats.cell(row=r, column=1)
            cell.fill = region_fill
            cell.font = region_font

    # 합계 행 추가
    ws_stats.row_dimensions[row_idx].height = 22
    cell_total_label = ws_stats.cell(row=row_idx, column=1, value="전체 합계")
    ws_stats.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
    
    cell_total_raw = ws_stats.cell(row=row_idx, column=3, value=f"=SUM(C2:C{row_idx-1})")
    cell_total_uniq = ws_stats.cell(row=row_idx, column=4, value=f"=SUM(D2:D{row_idx-1})")

    # 합계 행 스타일 적용
    for col in range(1, 5):
        c = ws_stats.cell(row=row_idx, column=col)
        c.font = summary_font
        c.fill = summary_fill
        c.border = summary_border
        if col >= 3:
            c.alignment = data_align_right
        else:
            c.alignment = data_align_center

    # --- 시트 2: 고유 회원 리스트 생성 ---
    headers_list = [
        "연번", "회원ID", "이름", "성별", "나이", 
        "회원 구분", "최종 점검일", "지역", "지점", 
        "종합점수", "신체나이", "뇌나이", "마음나이", 
        "얼굴나이", "종합건강나이", "대표 취약코드"
    ]
    
    ws_list.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers_list, 1):
        cell = ws_list.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 고유 회원 데이터 정렬 (지역명 -> 지점명 -> 이름 순)
    def sort_key(member):
        r_id = member.get('regionId', '')
        b_id = member.get('branchId', '')
        r_name = regions.get(r_id, {}).get('name', r_id)
        b_name = branches.get(b_id, {}).get('name', b_id)
        name = member.get('name', '')
        return (r_name, b_name, name)

    unique_members.sort(key=sort_key)

    for seq, m in enumerate(unique_members, 1):
        r_num = seq + 1
        ws_list.row_dimensions[r_num].height = 18

        r_id = m.get('regionId', '')
        b_id = m.get('branchId', '')
        region_name = regions.get(r_id, {}).get('name', r_id.replace('region_', '') if r_id else '미지정')
        branch_name = branches.get(b_id, {}).get('name', '미지정')

        report = m.get('report', {})
        ui = report.get('userInfo', {}) if isinstance(report, dict) else {}
        
        name = m.get('name', '').strip()
        phone = ui.get('phone', '') if isinstance(ui, dict) else ''
        gender = ui.get('gender', '')
        # 한글 성별 변환
        gender_kr = '여성' if gender == 'female' else ('남성' if gender == 'male' else gender)
        age = ui.get('age', '')
        
        m_type = ui.get('memberType', '')
        m_type_kr = '신규' if m_type == 'new' else ('기존' if m_type == 'existing' else m_type)
        
        test_date = m.get('lastTestDate', '')
        if test_date and len(test_date) >= 10:
            test_date = test_date[:10]  # YYYY-MM-DD 만 추출

        # 나이 및 점수 지표 추출
        overall_score = report.get('overallScore', '')
        physical_age = report.get('physicalAge', '')
        brain_age = report.get('brainAge', '')
        mind_age = report.get('mindAge', '')
        face_age = report.get('faceAgeEstimate', '')
        comprehensive_age = report.get('comprehensiveAge', '')
        
        # 취약 코드 파악 (가장 점수가 낮은 세븐코드)
        weakest_code_idx = report.get('sevenCodeAnalysis', {}).get('weakestCode')
        if not weakest_code_idx and 'sevenCodeAnalysis' in report:
            # 직접 계산
            min_score = 999
            min_code = None
            for k, v in report.get('sevenCodeAnalysis', {}).items():
                if k.startswith('code') and isinstance(v, dict) and 'score' in v:
                    score = v['score']
                    if score < min_score:
                        min_score = score
                        min_code = k
            if min_code:
                weakest_code_idx = min_code.replace('code', '')
        
        weakest_code_str = f"{weakest_code_idx}코드" if weakest_code_idx else "N/A"

        # 데이터 쓰기
        row_vals = [
            seq,
            m.get('id', ''),
            name,
            gender_kr,
            age,
            m_type_kr,
            test_date,
            region_name,
            branch_name,
            overall_score,
            physical_age,
            brain_age,
            mind_age,
            face_age,
            comprehensive_age,
            weakest_code_str
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws_list.cell(row=r_num, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            # 컬럼별 정렬 설정 (컬럼 수 축소에 따른 조정)
            if col_idx in [1, 2, 4, 6, 7, 8, 9, 16]: # 연번, ID, 성별, 구분, 날짜, 지역, 지점, 취약코드
                cell.alignment = data_align_center
            elif col_idx == 3: # 이름
                cell.alignment = data_align_left
            else: # 나이 및 점수 등 숫자형 데이터 (5: 나이, 10~15: 각종 점수 및 나이 지표)
                cell.alignment = data_align_right
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0'

    # --- 두 시트 모두 컬럼 너비 자동 조절 ---
    for ws in [ws_stats, ws_list]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                # 한글은 길이 계산 시 2글자로 판정하도록 보정
                byte_len = 0
                for char in val_str:
                    if ord(char) > 127:
                        byte_len += 2
                    else:
                        byte_len += 1
                if byte_len > max_len:
                    max_len = byte_len
            col_letter = get_column_letter(col[0].column)
            # 패딩 부여
            ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

    # 6. 파일 저장
    output_path = 'BTC_코드맵_점검_통계보고서.xlsx'
    wb.save(output_path)
    print(f"\n==================================================")
    print(f"[완료] 보고서 생성 완료: {output_path}")
    print(f"==================================================")

if __name__ == '__main__':
    build_excel_report()
